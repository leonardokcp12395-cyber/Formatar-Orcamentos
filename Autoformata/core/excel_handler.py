import pandas as pd
import openpyxl
import os
import shutil
import re
import math
from copy import copy
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import range_boundaries, get_column_letter
from utils.logger import Logger

class OrcamentoEngine:
    def __init__(self, config):
        self.output_dir = "Output"
        if not os.path.exists(self.output_dir): os.makedirs(self.output_dir)
        # Variáveis internas para compartilhar entre os métodos
        self.wb_out = None
        self.ws_out = None
        self.wb_src = None
        self.ws_src = None
        self.info = {}
        self.mapa_colunas = {}
        self.FMT_CONTABIL = '_("R$"* #,##0.00_);_("R$"* (#,##0.00);_("R$"* "-"??_);_(@_)'

    def gerar_excel_final(self, linhas_aprovadas, modelo_path, mapa_colunas, info):
        Logger.info(">>> ENGINE V32: ESTRUTURA REFATORADA <<<")
        self.info = info
        self.mapa_colunas = mapa_colunas
        
        try:
            save_path = self._preparar_arquivo(modelo_path)
            
            # 1. Processa o Cabeçalho
            self._processar_cabecalho()

            # 2. Identifica onde começar a escrever
            start_row = self._encontrar_inicio_tabela()
            
            # 3. Processa os Itens (o coração do programa)
            current_row, mapa_linhas = self._processar_itens(linhas_aprovadas, start_row)

            # 4. Insere os subtotais das etapas
            self._inserir_formulas_totais(mapa_linhas)

            # 5. Processa o Rodapé (Totais finais e cópia do modelo)
            self._processar_rodape(current_row, start_row)
            
            # Salva e fecha
            self.wb_out.save(save_path)
            self.wb_src.close()
            Logger.info(f"✅ Concluído: {save_path}")
            return True, save_path, {}

        except Exception as e:
            Logger.error(f"Erro Engine: {e}")
            import traceback
            traceback.print_exc()
            if self.wb_src: self.wb_src.close()
            return False, str(e), {}

    # ================= MÉTODOS AUXILIARES PRINCIPAIS =================

    def _preparar_arquivo(self, modelo_path):
        nome_arq = self.info.get('nome_arquivo', 'Orcamento').strip()
        nome_arq = re.sub(r'[<>:"/\\|?*]', '', nome_arq)
        save_path = os.path.join(self.output_dir, f"{nome_arq}.xlsx")
        shutil.copy(modelo_path, save_path)
        
        self.wb_out = openpyxl.load_workbook(save_path)
        self.ws_out = self.wb_out.active
        self.wb_src = openpyxl.load_workbook(modelo_path)
        self.ws_src = self.wb_src.active
        return save_path

    def _processar_cabecalho(self):
        info = self.info
        self._write_cell('A8', f"CAMPUS: {info.get('campus', '')}", bold=True)
        self._write_cell('A9', f"SETOR:  {info.get('setor', '')}", bold=True)
        self._write_cell('A10', f"SERVIDOR: {info.get('servidor', '')}", bold=True)
        self._write_cell('A13', f"ORÇAMENTO ELABORADO POR: {info.get('elaborador', '')}", bold=True)
        self._write_cell('A14', f"ESTAGIÁRIO: {info.get('estagiario', '')}", bold=True)
        self._write_cell('C15', info.get('descricao_header', '').upper(), bold=False)
        dt = info.get('data', 'xx/xx/xxxx')
        self._write_cell('A18', f"DATA DE ELABORAÇÃO DO ORÇAMENTO: {dt} (VALIDADE: 45 DIAS)", bold=True)
        self._write_cell('E18', f"CÓDIGO ORÇAFASCIO:  {info.get('orcafascio', '')}", bold=True)
        self._write_cell('E21', f"NÚMERO DO PROCESSO:  {info.get('processo', '')}", bold=True)
        self._write_cell('D22', f"FISCAL DO SERVIÇO: {info.get('fiscal', '')}", bold=True)

    def _encontrar_inicio_tabela(self):
        start_row = 15
        for r in range(1, 50):
            val = str(self.ws_out.cell(r, 4).value).upper()
            if 'DESCRIÇÃO' in val or 'DISCRIMINAÇÃO' in val:
                start_row = r + 1
                break
        return start_row

    def _processar_itens(self, linhas_aprovadas, start_row):
        FMT_NUM = '0.00'
        FMT_MOEDA = '"R$ "#,##0.00'
        
        calc_mode = self.info.get('calc_mode', 'EXACT')
        altura_base = self.info.get('altura_linha', 24.75)
        
        cols = {k: self.mapa_colunas[k] for k in ["ITEM","CODIGO","BANCO","DESCRICAO","UNID","QUANT","UNIT"]}
        mapa_linhas_escritas = []
        current_row = start_row

        for row_data in linhas_aprovadas:
            self._limpar_mesclagem_linha(current_row)
            nivel = row_data["_NIVEL_FORCADO"]
            
            self._safe_write(current_row, 1, row_data.get(cols["ITEM"], ''))
            self._safe_write(current_row, 2, row_data.get(cols["CODIGO"], ''))
            self._safe_write(current_row, 3, row_data.get(cols["BANCO"], ''))
            self._safe_write(current_row, 4, row_data.get(cols["DESCRICAO"], ''))

            if nivel == "ITEM":
                self._safe_write(current_row, 5, row_data.get(cols["UNID"], ''))
                
                qtd_raw = self._parse_num(row_data.get(cols["QUANT"]))
                unit_raw = self._parse_num(row_data.get(cols["UNIT"]))
                
                qtd_final = self._aplicar_precisao(qtd_raw, calc_mode)
                unit_final = self._aplicar_precisao(unit_raw, calc_mode)

                if qtd_final is not None: self._safe_write(current_row, 6, qtd_final, FMT_NUM)
                if unit_final is not None: self._safe_write(current_row, 7, unit_final, FMT_MOEDA)
                    
                # APLICAÇÃO DO ROUNDDOWN NA FÓRMULA (V31)
                self._safe_write(current_row, 8, f"=ROUNDDOWN(F{current_row}*G{current_row}, 2)", self.FMT_CONTABIL)
                mapa_linhas_escritas.append({'row': current_row, 'nivel': 'ITEM'})
            else:
                mapa_linhas_escritas.append({'row': current_row, 'nivel': nivel})

            self._aplicar_estilo_hierarquico(current_row, nivel)
            self._ajustar_altura_linha(current_row, row_data.get(cols["DESCRICAO"],''), altura_base)
            
            current_row += 1
            
        return current_row, mapa_linhas_escritas

    def _processar_rodape(self, current_row, start_row):
        ultima_linha_dados = current_row - 1
        max_row_planilha = max(self.ws_out.max_row, 200)
        self._limpar_area_total(current_row, max_row_planilha)

        target_start_row = current_row 
        self._copiar_bloco_excel(26, 51, target_start_row)
        
        bdi_val = self.info.get("bdi", 0.0)
        fator_desconto = 0.19 if abs(bdi_val - 0.2882) < 0.001 else 0.0601

        r1, r2, r3, r4, r5 = [target_start_row + i for i in range(5)]
        font_bold = Font(name="Arial", bold=True, size=10)

        self.ws_out.cell(r1, 8).value = f"=SUBTOTAL(9, H{start_row}:H{ultima_linha_dados})"
        self.ws_out.cell(r2, 8).value = f"=H{r1}*{bdi_val}"
        self.ws_out.cell(r3, 8).value = f"=H{r1}+H{r2}"
        self.ws_out.cell(r4, 8).value = f"=H{r3}*{fator_desconto}" 
        self.ws_out.cell(r5, 8).value = f"=H{r3}-H{r4}"

        for r in [r1, r2, r3, r4, r5]:
            c = self.ws_out.cell(r, 8)
            c.number_format = self.FMT_CONTABIL
            c.font = font_bold

    # ================= MÉTODOS UTILITÁRIOS =================

    def _write_cell(self, coord, text, bold=True):
        """Escreve em uma célula específica (usada no cabeçalho)"""
        cell = self.ws_out[coord]
        if isinstance(cell, MergedCell):
            for merged in self.ws_out.merged_cells.ranges:
                if cell.coordinate in merged:
                    cell = self.ws_out.cell(merged.min_row, merged.min_col)
                    break
        cell.value = text
        current_font = copy(cell.font)
        new_font = Font(name=current_font.name, size=current_font.size, bold=bold, color=current_font.color)
        cell.font = new_font

    def _safe_write(self, row, col, value, number_format=None):
        """Escreve em uma célula por linha/coluna, lidando com mesclagens"""
        cell = self.ws_out.cell(row, col)
        if isinstance(cell, MergedCell):
            for merged in list(self.ws_out.merged_cells.ranges):
                if cell.coordinate in merged: self.ws_out.unmerge_cells(str(merged)); break
            cell = self.ws_out.cell(row, col)
        cell.value = value
        if number_format: cell.number_format = number_format

    def _ajustar_altura_linha(self, row, desc_txt, altura_base):
        desc_txt = str(desc_txt)
        num_chars = len(desc_txt)
        linhas_estimadas = max(1, math.ceil(num_chars / 85))
        if linhas_estimadas == 1:
            altura_final = altura_base
        else:
            altura_final = max(altura_base, linhas_estimadas * 15)
        self.ws_out.row_dimensions[row].height = altura_final

    def _aplicar_precisao(self, valor, modo):
        if valor is None: return None
        try:
            val_float = float(valor)
        except:
            return valor

        if modo == "TRUNC":
            # MÉTODO STRING SLICING (INFALÍVEL)
            s = f"{val_float:.10f}"
            if '.' in s:
                int_part, dec_part = s.split('.')
                dec_final = dec_part[:2]
                return float(f"{int_part}.{dec_final}")
            return val_float
        elif modo == "ROUND":
            return round(val_float, 2)
        else:
            return val_float

    def _limpar_area_total(self, row_inicio, row_fim):
        for merged in list(self.ws_out.merged_cells.ranges):
            if merged.max_row >= row_inicio:
                try: self.ws_out.unmerge_cells(str(merged))
                except: pass
        rows_to_delete = row_fim - row_inicio + 1
        if rows_to_delete > 0:
            self.ws_out.delete_rows(row_inicio, rows_to_delete)

    def _limpar_mesclagem_linha(self, row):
        for merged in list(self.ws_out.merged_cells.ranges):
            if row >= merged.min_row and row <= merged.max_row:
                try: self.ws_out.unmerge_cells(str(merged))
                except: pass

    def _copiar_bloco_excel(self, r_ini, r_fim, r_tgt_ini):
        offset = r_tgt_ini - r_ini
        for row in range(r_ini, r_fim + 1):
            tgt_row = row + offset
            self.ws_out.row_dimensions[tgt_row].height = self.ws_src.row_dimensions[row].height
            for col in range(1, self.ws_src.max_column + 1):
                cell_src = self.ws_src.cell(row, col)
                cell_tgt = self.ws_out.cell(tgt_row, col)
                cell_tgt.value = cell_src.value
                if cell_src.has_style:
                    cell_tgt.font = copy(cell_src.font)
                    cell_tgt.border = copy(cell_src.border)
                    cell_tgt.fill = copy(cell_src.fill)
                    cell_tgt.number_format = cell_src.number_format
                    cell_tgt.alignment = copy(cell_src.alignment)

        for merged in self.ws_src.merged_cells.ranges:
            min_c, min_r, max_c, max_r = range_boundaries(str(merged))
            if min_r >= r_ini and max_r <= r_fim:
                new_min_r = min_r + offset
                new_max_r = max_r + offset
                coord_start = f"{get_column_letter(min_c)}{new_min_r}"
                coord_end = f"{get_column_letter(max_c)}{new_max_r}"
                try: self.ws_out.merge_cells(f"{coord_start}:{coord_end}")
                except ValueError: pass

    def _inserir_formulas_totais(self, mapa):
        try:
            total_rows = len(mapa)
            peso = {"N1": 1, "N2": 2, "N3": 3, "ITEM": 4}
            for i, atual in enumerate(mapa):
                if atual['nivel'] == "ITEM": continue
                row_pai = atual['row']
                nivel_pai = peso.get(atual['nivel'], 1)
                idx_fim = i
                for j in range(i + 1, total_rows):
                    prox = mapa[j]
                    if peso.get(prox['nivel'], 4) <= nivel_pai: break
                    idx_fim = j
                if idx_fim > i:
                    r_ini = mapa[i+1]['row']
                    r_fim = mapa[idx_fim]['row']
                    cell = self.ws_out.cell(row_pai, 8)
                    cell.value = f"=SUBTOTAL(9, H{r_ini}:H{r_fim})"
                    cell.number_format = self.FMT_CONTABIL
                    cell.font = Font(bold=True)
        except Exception as e: Logger.error(f"Erro totais: {e}")

    def _aplicar_estilo_hierarquico(self, row, nivel):
        paleta = {
            "N1": {"bg": "9BC2E6", "bold": True, "size": 11},
            "N2": {"bg": "BDD7EE", "bold": True, "size": 11},
            "N3": {"bg": "DDEBF7", "bold": True, "size": 11},
            "ITEM": {"bg": "FFFFFF", "bold": False, "size": 10}
        }
        estilo = paleta.get(nivel, paleta["ITEM"])
        fill = PatternFill("solid", fgColor=estilo["bg"])
        font = Font(name="Arial", bold=estilo["bold"], size=estilo["size"])
        side = Side(style="thin")
        border = Border(left=side, right=side, top=side, bottom=side)
        for c in range(1, 9):
            cell = self.ws_out.cell(row, c)
            if isinstance(cell, MergedCell): continue
            cell.fill = fill
            cell.font = font
            cell.border = border
            h = "center"
            if c == 4: h = "left"
            if c == 8: h = "right"
            cell.alignment = Alignment(horizontal=h, vertical="center", wrap_text=(c==4))

    def _parse_num(self, val):
        if pd.isna(val): return None
        try:
            s = str(val).replace('R$', '').strip()
            if ',' in s and '.' in s: s = s.replace('.', '').replace(',', '.')
            elif ',' in s: s = s.replace(',', '.')
            return float(s)
        except: return None