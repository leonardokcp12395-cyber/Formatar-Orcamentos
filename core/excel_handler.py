import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from datetime import datetime
import os
import shutil
import unicodedata
from utils.logger import Logger

class OrcamentoEngine:
    def __init__(self, config):
        self.config = config
        self.output_dir = self.config.get('output', {}).get('dir', 'orcamentos_gerados')
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

        # Definição de estilos
        # Cores aproximadas do Excel (Azul claro degradê)
        self.cores = {
            'N1': PatternFill(start_color="4472C4", fill_type="solid"), # Azul Escuro (Título Principal)
            'N2': PatternFill(start_color="8EA9DB", fill_type="solid"), # Azul Médio
            'N3': PatternFill(start_color="D9E1F2", fill_type="solid"), # Azul Claro
            'TITULO': PatternFill(start_color="D9E1F2", fill_type="solid") # Genérico
        }

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        self.font_bold = Font(bold=True)

    def processar_orcamento(self, arquivo_limpo, modelo_path, dados_projeto, mapa_niveis, intervalo_linhas):
        """
        Processa o orçamento com lógica aprimorada de níveis e rodapé.
        """
        Logger.titulo("INICIANDO ENGINE DE ORÇAMENTO (V2)")

        try:
            linha_ini, linha_fim = intervalo_linhas
            Logger.info(f"Intervalo: {linha_ini} a {linha_fim}")

            # 1. Detectar Cabeçalho (usando 100 primeiras linhas)
            df_search = pd.read_excel(arquivo_limpo, header=None, nrows=100)
            col_map, header_row = self._detectar_colunas(df_search)

            Logger.info(f"Mapa: {col_map} (Linha {header_row})")

            # 2. Ler Dados
            skip = linha_ini - 1
            qtd = linha_fim - linha_ini + 1

            df_dados = pd.read_excel(arquivo_limpo, header=None, skiprows=skip, nrows=qtd)
            df_dados = df_dados.dropna(how='all')

            if df_dados.empty:
                return False, "Nenhum dado encontrado.", {}

            # 3. Preparar Arquivo
            filename = f"ORC_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path = os.path.join(self.output_dir, filename)
            shutil.copy(modelo_path, save_path)

            wb = openpyxl.load_workbook(save_path)
            ws = wb.active

            # 4. Preencher Cabeçalho Fixo
            self._preencher_cabecalho(ws, dados_projeto)

            # 5. Inserir Itens
            linha_base = self.config.get('excel', {}).get('linha_inicial_modelo', 25)
            current_row = linha_base
            itens_processados = 0

            for idx, row in df_dados.iterrows():
                # Extrair valores
                val_item = self._get_val(row, col_map['item'])
                val_desc = self._get_val(row, col_map['desc'])
                val_unid = self._get_val(row, col_map['unid'])
                val_quant = self._get_float(row, col_map['quant'])
                val_unit = self._get_float(row, col_map['unit'])
                val_cod = self._get_val(row, col_map['cod'])

                # Inferir nível: Se tem Preço/Quant, é ITEM. Senão, é TÍTULO.
                # Prioridade: Mapa manual > Inferência
                nivel = mapa_niveis.get(idx)
                if not nivel:
                    if val_unit is not None or val_quant is not None:
                        nivel = 'ITEM'
                    else:
                        # Tenta inferir profundidade pelos pontos (1. vs 1.1.)
                        nivel = self._inferir_titulo(val_item)

                ws.insert_rows(current_row)

                # Colunas: A=Item, B=Cod, C=Banco(Vazio), D=Desc, E=Und, F=Quant, G=Unit, H=Total
                ws.cell(row=current_row, column=1, value=val_item)
                ws.cell(row=current_row, column=2, value=val_cod)
                ws.cell(row=current_row, column=4, value=val_desc)

                if nivel == 'ITEM':
                    ws.cell(row=current_row, column=5, value=val_unid)

                    if val_quant is not None:
                        ws.cell(row=current_row, column=6, value=val_quant)
                        ws.cell(row=current_row, column=6).number_format = '#,##0.00'

                    if val_unit is not None:
                        ws.cell(row=current_row, column=7, value=val_unit)
                        ws.cell(row=current_row, column=7).number_format = '#,##0.00'

                    if val_quant is not None and val_unit is not None:
                         ws.cell(row=current_row, column=8, value=f"=F{current_row}*G{current_row}")
                         ws.cell(row=current_row, column=8).number_format = '#,##0.00'

                self._aplicar_estilo(ws, current_row, nivel)
                current_row += 1
                itens_processados += 1

            # 6. Atualizar Rodapé (Preservando linhas existentes)
            self._atualizar_rodape(ws, current_row, linha_base, dados_projeto)

            wb.save(save_path)
            return True, save_path, {'itens': itens_processados}

        except Exception as e:
            Logger.error(f"Erro no Engine: {e}")
            import traceback
            Logger.error(traceback.format_exc())
            return False, str(e), {}

    def _detectar_colunas(self, df):
        col_map = {'item': -1, 'desc': -1, 'unid': -1, 'quant': -1, 'unit': -1, 'cod': -1}
        header_row = 0

        for r_idx, row in df.iterrows():
            row_vals = [self._normalize_str(v) for v in row.values]
            if any("ITEM" in v for v in row_vals) and \
               (any("DESCRI" in v for v in row_vals) or any("DISCRIMINA" in v for v in row_vals)):
                header_row = r_idx
                for c_idx, val in enumerate(row_vals):
                    if self._match(val, ['ITEM', 'IT', 'NUM']): col_map['item'] = c_idx
                    elif self._match(val, ['DESCRI', 'DISCRIMINA']): col_map['desc'] = c_idx
                    elif self._match(val, ['UNID', 'UND', 'UNIDADE']): col_map['unid'] = c_idx
                    elif self._match(val, ['QUANT', 'QTD', 'QTDE']): col_map['quant'] = c_idx
                    elif self._match(val, ['UNIT', 'PRECO', 'VALOR']): col_map['unit'] = c_idx
                    elif self._match(val, ['COD', 'SINAPI']): col_map['cod'] = c_idx
                break

        # Fallbacks posicionais
        if col_map['item'] == -1: col_map['item'] = 0
        if col_map['desc'] == -1: col_map['desc'] = 1

        return col_map, header_row

    def _inferir_titulo(self, val_item):
        if not val_item: return 'N1'
        dots = val_item.count('.')
        if dots == 0: return 'N1'
        if dots == 1: return 'N2'
        return 'N3'

    def _preencher_cabecalho(self, ws, dados):
        try:
            if dados.get('obra'): ws['C8'] = dados['obra']
            if dados.get('local'): ws['C9'] = dados['local']
            if dados.get('email'): ws['H10'] = dados['email']
            ws['C17'] = datetime.now().strftime("%d/%m/%Y")
        except: pass

    def _atualizar_rodape(self, ws, current_row, linha_base, dados):
        # Procura "Total sem BDI" nas linhas seguintes
        # Limite de busca: 50 linhas para baixo
        for r in range(current_row, current_row + 50):
            val_a = ws.cell(row=r, column=1).value
            if val_a and isinstance(val_a, str) and "Total sem BDI" in val_a:

                # Encontrou o início do rodapé
                last_item = current_row - 1
                if last_item < linha_base: return # Nada inserido

                # Linha 1: Total sem BDI
                ws.cell(row=r, column=8, value=f"=SUM(H{linha_base}:H{last_item})")
                ws.cell(row=r, column=8).number_format = '#,##0.00'

                # Linha 2: BDI
                # Assume que é a próxima linha. Pode checar o texto da col A também.
                bdi_val = self._parse_bdi(dados.get('bdi'))
                ws.cell(row=r+1, column=8, value=f"=H{r}*{bdi_val}")
                ws.cell(row=r+1, column=8).number_format = '#,##0.00'

                # Linha 3: Total Geral
                ws.cell(row=r+2, column=8, value=f"=H{r}+H{r+1}")
                ws.cell(row=r+2, column=8).number_format = '#,##0.00'

                # Linha 4: Desconto (Se existir) e Linha 5: Total com Desconto
                # O Modelo tem "Valor do Desconto em ATA (19%)" na linha r+3 (index 3 relativo a r)
                # Vamos checar se existe texto na A
                val_desc = ws.cell(row=r+3, column=1).value
                if val_desc and "Desconto" in str(val_desc):
                    # Tenta extrair percentual do texto "ATA (19%)" ou similar, ou deixa 0
                    # Vou assumir 0 se não for parametrizado, ou tentar manter fórmula se já existir?
                    # Como inserimos linhas, fórmulas antigas podem ter quebrado se referenciam células fixas acima.
                    # Mas fórmulas de rodapé geralmente referenciam o total acima.

                    # Vamos colocar 0 por padrão se não tivermos info de desconto
                    # Ou tentar extrair do texto? Ex: "19%"
                    desc_pct = 0.0
                    if "%" in str(val_desc):
                        try:
                            import re
                            nums = re.findall(r'(\d+[.,]?\d*)%', str(val_desc))
                            if nums:
                                desc_pct = float(nums[0].replace(',', '.')) / 100
                        except: pass

                    ws.cell(row=r+3, column=8, value=f"=H{r+2}*{desc_pct}") # Aplica sobre o Geral? Ou Sem BDI? Geralmente sobre Geral.
                    ws.cell(row=r+3, column=8).number_format = '#,##0.00'

                    # Total Final
                    ws.cell(row=r+4, column=8, value=f"=H{r+2}-H{r+3}")
                    ws.cell(row=r+4, column=8).number_format = '#,##0.00'

                break

    def _match(self, val, keys):
        v = str(val).upper()
        return any(k in v for k in keys)

    def _normalize_str(self, text):
        if pd.isna(text): return ""
        t = str(text).upper().strip()
        return ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')

    def _get_val(self, row, idx):
        if idx < 0 or idx >= len(row): return ""
        v = row.iloc[idx]
        return "" if pd.isna(v) else str(v).strip()

    def _get_float(self, row, idx):
        if idx < 0 or idx >= len(row): return None
        v = row.iloc[idx]
        if pd.isna(v): return None
        if isinstance(v, (int, float)): return v
        try:
            s = str(v).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
            return float(s)
        except: return None

    def _parse_bdi(self, val):
        try:
            if isinstance(val, str): val = val.replace(',', '.')
            return float(val) / 100
        except: return 0.0

    def _aplicar_estilo(self, ws, row, nivel):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = self.border

            # Alinhamento
            if col == 4: # Descrição
                cell.alignment = self.align_left
            else:
                cell.alignment = self.align_center # Padrão centralizado

            # Cores apenas para Níveis de Título
            if nivel in self.cores and nivel != 'ITEM':
                cell.fill = self.cores[nivel]
                cell.font = self.font_bold
                # Títulos geralmente centralizados ou esquerda? Descrição esquerda.
                # Se for título, talvez tudo negrito.

            # Ajuste fino: Item sempre negrito? N1, N2, N3 sim. ITEM normal.
            if nivel == 'ITEM':
                cell.font = Font(bold=False)
            else:
                cell.font = self.font_bold
